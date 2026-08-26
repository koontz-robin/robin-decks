#!/usr/bin/env python3
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict, Counter
from difflib import SequenceMatcher
import json, glob, re, html, csv, datetime

ROOT=Path('/home/openclaw/.openclaw/workspace')
XLSX=Path('/home/openclaw/.openclaw/media/inbound/Navigate26_Pre_Conf_Reg_List---ccdef9e6-ac17-4052-aa52-4834921b1119.xlsx')
OUT_JSON=ROOT/'navigate26_company_owner_dashboard.json'
OUT_CSV=ROOT/'navigate26_company_owner_dashboard.csv'
OUT_HTML=ROOT/'navigate26-company-owner-dashboard.html'

SUFFIX_RE=re.compile(r'\b(incorporated|inc|llc|l\.l\.c|ltd|co|company|corp|corporation|limited|pllc|lp|llp|pc)\b', re.I)
STOP={'the','and','of','communications','communication','telecommunications','telecom','technology','technologies','solutions','systems','services','service','internet','network','networks','wireless','broadband','business','telephone','phone','group','partners','association','coop','cooperative','rural','membership'}

def norm(s):
    s=str(s or '').strip().lower()
    s=s.replace('&',' and ')
    s=re.sub(r'\bdba\b.*$','',s)
    s=SUFFIX_RE.sub(' ',s)
    s=re.sub(r'[^a-z0-9]+',' ',s)
    s=re.sub(r'\s+',' ',s).strip()
    return s

def key_tokens(s):
    return [t for t in norm(s).split() if t not in STOP and len(t)>1]

def owner_name(v):
    if isinstance(v, dict): return v.get('Name') or v.get('name') or ''
    return v or ''

candidates=defaultdict(list)
def add_candidate(name, owner, source, priority, meta=None):
    name=str(name or '').strip(); owner=str(owner or '').strip()
    if not name or not owner: return
    candidates[norm(name)].append({'account':name,'owner':owner,'source':source,'priority':priority,'meta':meta or {}})

# True account-owner cache
try:
    j=json.load(open(ROOT/'sf_account_owners.json'))
    for r in j.get('records',[]): add_candidate(r.get('Name'), owner_name(r.get('Owner')), 'sf_account_owners.json Account.Owner', 100, {'id': (r.get('attributes') or {}).get('url','').split('/')[-1]})
except Exception: pass

# Account-ish exports
for fn in ['sf_smc_reengagement_accounts.json','sf_msp_psa_clients.json','sf_msp_psa_web_clients.json']:
    p=ROOT/fn
    if not p.exists(): continue
    try: data=json.load(open(p))
    except Exception: continue
    rows=data.get('rows',[]) if isinstance(data,dict) else data if isinstance(data,list) else []
    for r in rows:
        if not isinstance(r,dict): continue
        acct=r.get('account') or r.get('Account') or r.get('name') or r.get('Name')
        # account_owner beats salesperson; salesperson is account-ish fallback in these exports
        if r.get('account_owner'): add_candidate(acct, r.get('account_owner'), fn+' account_owner', 95, {'url':r.get('account_url','')})
        if r.get('owner'): add_candidate(acct, r.get('owner'), fn+' owner', 85, {'url':r.get('account_url','')})
        if r.get('salesperson'): add_candidate(acct, r.get('salesperson'), fn+' salesperson', 80, {'url':r.get('account_url','')})


# Generic local Salesforce-derived dashboard/cache exports with explicit account_owner fields.
def walk_json(obj, fn):
    if isinstance(obj, dict):
        acct=obj.get('account') or obj.get('Account') or obj.get('account_name') or obj.get('AccountName') or obj.get('company') or obj.get('Company') or obj.get('Name')
        if isinstance(acct, dict): acct=acct.get('Name')
        if obj.get('account_owner'):
            add_candidate(acct, obj.get('account_owner'), fn+' account_owner', 92, {'url':obj.get('account_url','')})
        elif isinstance(obj.get('Owner'), dict) and obj['Owner'].get('Name') and (acct or obj.get('Name')):
            # only if the object appears account-like, not every opportunity-like generic object
            typ=((obj.get('attributes') or {}).get('type') or '').lower()
            if typ == 'account': add_candidate(acct or obj.get('Name'), obj['Owner'].get('Name'), fn+' Account.Owner', 100, {})
        for v in obj.values(): walk_json(v, fn)
    elif isinstance(obj, list):
        for v in obj: walk_json(v, fn)

for f in glob.glob(str(ROOT/'*.json')):
    fn=Path(f).name
    if fn.startswith('node_modules'): continue
    try: data=json.load(open(f))
    except Exception: continue
    walk_json(data, fn)

# Opportunity exports fallback: useful when account-owner cache is stale/missing, but flag as lower confidence.
for f in glob.glob(str(ROOT/'sf_*.json')):
    fn=Path(f).name
    if fn in {'sf_account_owners.json','sf_smc_reengagement_accounts.json','sf_msp_psa_clients.json','sf_msp_psa_web_clients.json'}: continue
    try: data=json.load(open(f))
    except Exception: continue
    if isinstance(data,dict):
        rows=[]
        if isinstance(data.get('records'),list): rows=data['records']
        elif isinstance(data.get('opportunities'),list): rows=data['opportunities']
        elif isinstance(data.get('rows'),list): rows=data['rows']
    elif isinstance(data,list): rows=data
    else: rows=[]
    for r in rows:
        if not isinstance(r,dict): continue
        acct=r.get('Account') or r.get('account')
        if isinstance(acct,dict): acct=acct.get('Name')
        owner=owner_name(r.get('Owner') or r.get('owner'))
        if acct and owner: add_candidate(acct, owner, fn+' opportunity owner fallback', 60, {'opp':r.get('Name','')})

# Collapse candidate names to best owner per SF normalized name
sf_names=[]
for k, vals in candidates.items():
    # best priority, then frequency by owner at that priority
    maxp=max(v['priority'] for v in vals)
    owners=[v['owner'] for v in vals if v['priority']==maxp]
    owner=Counter(owners).most_common(1)[0][0]
    ex=next(v for v in vals if v['priority']==maxp and v['owner']==owner)
    sf_names.append({'norm':k,'account':ex['account'],'owner':owner,'source':ex['source'],'priority':maxp,'alternates':vals})

# Read registration list
wb=load_workbook(XLSX, read_only=True, data_only=True)
ws=wb.active
rows=list(ws.iter_rows(values_only=True))
header=[str(x or '').strip() for x in rows[0]]
idx={h:i for i,h in enumerate(header)}
records=[]
for row in rows[1:]:
    rec={h:(row[i] if i<len(row) else '') for h,i in idx.items()}
    company=str(rec.get('Company') or '').strip()
    if not company: continue
    records.append(rec)

companies=defaultdict(lambda:{'company':'','attendees':0,'people':[]})
for r in records:
    c=str(r.get('Company') or '').strip()
    companies[norm(c)]['company']=c
    companies[norm(c)]['attendees']+=1
    companies[norm(c)]['people'].append({'first':str(r.get('First Name') or '').strip(),'last':str(r.get('Last Name') or '').strip(),'title':str(r.get('Job Title') or '').strip(),'email':str(r.get('Email Address') or '').strip()})

def match_company(company):
    n=norm(company)
    if n in candidates:
        # exact collapsed
        sf=next(x for x in sf_names if x['norm']==n)
        return {**sf,'match_type':'exact','score':1.0}
    toks=set(key_tokens(company))
    best=None
    for sf in sf_names:
        sn=sf['norm']
        # Only conservative fuzzy-match against account-level/account-ish exports.
        # Exact normalized matches are handled above. Here we only accept clear
        # prefix/contains variants like "Range" -> "Range Telecom" or
        # "Strata" -> "Strata Networks". No generic token roulette.
        if sf['priority'] < 80:
            continue
        ratio=SequenceMatcher(None,n,sn).ratio()
        stoks=set(key_tokens(sf['account']))
        inter=toks & stoks
        contain=(n and sn and (n in sn or sn in n) and min(len(n),len(sn))>=5)
        starts=(n and sn and (sn.startswith(n+' ') or n.startswith(sn+' ')) and min(len(n),len(sn))>=5)
        # Require either a clean prefix/contains relationship or very high string similarity
        # with the same first significant token.
        kt_company=key_tokens(company)
        if len(kt_company) < 2:
            continue
        first_company=kt_company[:1]
        first_sf=key_tokens(sf['account'])[:1]
        first_ok=bool(first_company and first_sf and first_company[0]==first_sf[0])
        score=max(ratio, .94 if starts else 0, .90 if contain else 0)
        if (starts or (contain and first_ok) or (ratio>=0.95 and first_ok and len(inter)>=2)) and (best is None or score>best['score']):
            best={**sf,'match_type':'fuzzy','score':score}
    return best

company_rows=[]
for c in companies.values():
    m=match_company(c['company'])
    if m:
        owner=m['owner']; status='Matched' if m['priority']>=80 else 'SF opp-owner fallback'; confidence='High' if m['priority']>=95 and m['score']>=.95 else 'Medium' if m['priority']>=80 else 'Low'
        sf_account=m['account']; source=m['source']; score=m['score']; mtype=m['match_type']
    else:
        owner='Unmatched'; status='Unmatched'; confidence='Needs SF lookup'; sf_account=''; source=''; score=0; mtype='none'
    company_rows.append({**c,'owner':owner,'status':status,'confidence':confidence,'sf_account':sf_account,'source':source,'score':round(score,3),'match_type':mtype})
company_rows.sort(key=lambda r:(r['owner']=='Unmatched', r['owner'], -r['attendees'], r['company'].lower()))

owner_groups=defaultdict(lambda:{'owner':'','companies':0,'attendees':0,'high':0,'medium':0,'low':0,'unmatched':0,'rows':[]})
for r in company_rows:
    g=owner_groups[r['owner']]; g['owner']=r['owner']; g['companies']+=1; g['attendees']+=r['attendees']; g['rows'].append(r)
    if r['confidence']=='High': g['high']+=1
    elif r['confidence']=='Medium': g['medium']+=1
    elif r['confidence']=='Low': g['low']+=1
    else: g['unmatched']+=1
owner_list=sorted(owner_groups.values(), key=lambda g:(g['owner']=='Unmatched', -g['attendees'], g['owner']))

out={'generated_at':datetime.datetime.utcnow().isoformat()+'Z','source_file':XLSX.name,'total_attendees':len(records),'unique_companies':len(company_rows),'matched_companies':sum(1 for r in company_rows if r['owner']!='Unmatched'),'unmatched_companies':sum(1 for r in company_rows if r['owner']=='Unmatched'),'owners':owner_list,'companies':company_rows}
OUT_JSON.write_text(json.dumps(out,indent=2))
with open(OUT_CSV,'w',newline='') as f:
    w=csv.DictWriter(f,fieldnames=['owner','company','attendees','status','confidence','sf_account','match_type','score','source'])
    w.writeheader(); w.writerows([{k:r.get(k,'') for k in w.fieldnames} for r in company_rows])

def esc(x): return html.escape(str(x or ''))
def owner_block(g):
    rows=''.join(f"<tr><td><strong>{esc(r['company'])}</strong><small>{esc(r['sf_account']) if r['sf_account'] and r['sf_account']!=r['company'] else ''}</small></td><td>{r['attendees']}</td><td><span class='badge {r['confidence'].lower().split()[0]}'>{esc(r['confidence'])}</span></td><td>{esc(r['source'])}</td></tr>" for r in g['rows'])
    return f"<section class='owner-card' data-owner='{esc(g['owner']).lower()}'><div class='owner-head'><div><h2>{esc(g['owner'])}</h2><p>{g['companies']} companies • {g['attendees']} attendees</p></div><div class='owner-total'>{g['attendees']}</div></div><table><thead><tr><th>Company</th><th>Attendees</th><th>Match</th><th>SF source</th></tr></thead><tbody>{rows}</tbody></table></section>"

top_rows=''.join(f"<tr><td>{i+1}</td><td>{esc(g['owner'])}</td><td>{g['attendees']}</td><td>{g['companies']}</td><td>{g['high']}</td><td>{g['medium']}</td><td>{g['low']}</td></tr>" for i,g in enumerate(owner_list[:20]) if g['owner']!='Unmatched')
blocks=''.join(owner_block(g) for g in owner_list)
html_doc=f"""<!doctype html><html lang='en'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>Navigate26 Registration List by Salesforce Owner</title><style>
:root{{--navy:#07131f;--blue:#2f80ed;--green:#54b948;--muted:#6b7787;--line:#dfe8f1;--paper:#f6f9fc;--orange:#f59e0b;--red:#ef4444}}*{{box-sizing:border-box}}body{{margin:0;font-family:Aptos,Inter,system-ui,-apple-system,Segoe UI,sans-serif;color:#102033;background:var(--paper)}}header{{background:radial-gradient(circle at 85% 5%,rgba(47,128,237,.38),transparent 34%),linear-gradient(125deg,#061019,#102d46 65%,#07131f);color:white;padding:34px clamp(20px,4vw,54px) 38px}}.top{{display:flex;justify-content:space-between;align-items:center;gap:24px}}.logo{{height:34px;object-fit:contain}}h1{{font-size:clamp(34px,5vw,64px);line-height:.98;margin:70px 0 18px;letter-spacing:-1.5px;max-width:980px}}header p{{color:rgba(255,255,255,.76);font-size:18px;line-height:1.45;max-width:900px}}.stats{{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:14px;margin-top:28px}}.stat{{background:rgba(255,255,255,.1);border:1px solid rgba(255,255,255,.18);border-radius:18px;padding:18px}}.stat b{{display:block;font-size:34px;color:#8bd5ff}}.stat span{{font-size:12px;text-transform:uppercase;font-weight:850;color:rgba(255,255,255,.72)}}main{{padding:28px clamp(16px,4vw,54px) 60px}}.toolbar{{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:20px}}input,select{{border:1px solid var(--line);border-radius:12px;padding:11px 12px;font:inherit;background:white;min-width:240px}}.summary,.owner-card{{background:white;border:1px solid var(--line);border-radius:20px;padding:20px;margin-bottom:18px;box-shadow:0 16px 42px rgba(7,19,31,.06)}}.summary h2,.owner-card h2{{margin:0 0 5px;font-size:24px}}.summary table,.owner-card table{{width:100%;border-collapse:collapse;margin-top:12px}}th{{text-align:left;color:#667085;font-size:12px;text-transform:uppercase;letter-spacing:.7px;border-bottom:1px solid var(--line);padding:10px}}td{{border-bottom:1px solid #eef3f7;padding:10px;vertical-align:top}}td small{{display:block;color:var(--muted);margin-top:3px}}.owner-head{{display:flex;justify-content:space-between;align-items:center;gap:16px}}.owner-head p{{margin:0;color:var(--muted)}}.owner-total{{font-size:34px;font-weight:900;color:var(--blue)}}.badge{{display:inline-block;border-radius:999px;padding:5px 9px;font-size:12px;font-weight:800;background:#eef6ff;color:#2563eb}}.badge.high{{background:#ecfdf3;color:#16803a}}.badge.medium{{background:#fff7ed;color:#c2410c}}.badge.low{{background:#fff1f2;color:#be123c}}.badge.needs{{background:#f2f4f7;color:#475467}}.note{{color:var(--muted);font-size:13px;line-height:1.45}}@media(max-width:800px){{.stats{{grid-template-columns:repeat(2,1fr)}}table{{font-size:13px}}th:nth-child(4),td:nth-child(4){{display:none}}}}
</style><script>function filt(){{const q=document.getElementById('q').value.toLowerCase();const o=document.getElementById('owner').value.toLowerCase();document.querySelectorAll('.owner-card').forEach(c=>{{const show=(!o||c.dataset.owner===o)&&(!q||c.innerText.toLowerCase().includes(q));c.style.display=show?'block':'none'}})}};</script></head><body><header><div class='top'><strong>Navigate26 pre-conference registration</strong><img class='logo' src='revio-logo-white.png' alt='Rev.io'></div><h1>Registration list broken out by Salesforce owner.</h1><p>Companies from the uploaded workbook were matched on the <strong>Company</strong> field against local Salesforce account-owner exports, with opportunity-owner fallback clearly flagged where a direct account-owner cache match was unavailable.</p><div class='stats'><div class='stat'><b>{len(records)}</b><span>registrants</span></div><div class='stat'><b>{len(company_rows)}</b><span>unique companies</span></div><div class='stat'><b>{out['matched_companies']}</b><span>matched companies</span></div><div class='stat'><b>{out['unmatched_companies']}</b><span>unmatched companies</span></div></div></header><main><div class='toolbar'><input id='q' oninput='filt()' placeholder='Search company / owner / source…'><select id='owner' onchange='filt()'><option value=''>All owners</option>{''.join(f"<option value='{esc(g['owner']).lower()}'>{esc(g['owner'])} — {g['attendees']}</option>" for g in owner_list)}</select></div><section class='summary'><h2>Owner leaderboard</h2><p class='note'>High = direct Account.Owner/account-owner source; Medium = account-ish Salesforce export; Low = opportunity-owner fallback from Salesforce exports. Unmatched companies need live Salesforce lookup or manual review.</p><table><thead><tr><th>#</th><th>Owner</th><th>Attendees</th><th>Companies</th><th>High</th><th>Med</th><th>Low</th></tr></thead><tbody>{top_rows}</tbody></table></section>{blocks}</main></body></html>"""
OUT_HTML.write_text(html_doc)
print(json.dumps({k:out[k] for k in ['total_attendees','unique_companies','matched_companies','unmatched_companies']}, indent=2))
print('top owners:', [(g['owner'],g['attendees'],g['companies']) for g in owner_list[:10]])
